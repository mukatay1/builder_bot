import os
from datetime import datetime

from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Border, Side

from utils.string_to_list import string_to_list

load_dotenv()

DEBUG = False

FOREMAN = string_to_list(os.getenv('FOREMAN'))
OKK_MAN = string_to_list(os.getenv('OKK_MAN'))
REPORT_MAN = string_to_list(os.getenv('REPORT_MAN'))
ADMIN = string_to_list(os.getenv('ADMIN'))

DEADLINE = "2024-11-15"

STATUS = [
    'Черновая электрика и черновая сантехника',
    'Стены',
    'Подоконник/откосы',
    'Укладка с\у и балкона',
    'Наливной пол',
    'Укладка ламината',
    'Обои',
    'Установка дверей',
    'Натяжной потолок',
    'Установка чистовой электрики и чистовой сантехники',
    'Покраска стен',
    'Клининг'
]

MAX_STATUS_NUMBER = len(STATUS)

red_fill = PatternFill(start_color='f2aca2', end_color='f2aca2', fill_type='solid')
green_fill = PatternFill(start_color='b1d492', end_color='b1d492', fill_type='solid')
yellow_fill = PatternFill(start_color='f7f3ac', end_color='f7f3ac', fill_type='solid')
blue_fill = PatternFill(start_color='b3c6e7', end_color='b3c6e7', fill_type='solid')
gray_fill = PatternFill(start_color='a4a6a3', end_color='a4a6a3', fill_type='solid')


black_border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

NUMBERS = [1, 2, 3, 4, 5]
lst = FOREMAN + OKK_MAN + REPORT_MAN + ADMIN
USERS = [user_id for user_id in lst if user_id]

plan_dates = [
    datetime(2024, 9, 15), datetime(2024, 10, 1), datetime(2024, 10, 1),
    datetime(2024, 10, 2), datetime(2024, 11, 15), datetime(2024, 11, 15),
    datetime(2024, 11, 15), datetime(2024, 11, 10), datetime(2024, 11, 10),
    datetime(2024, 10, 1), datetime(2024, 11, 1), datetime(2024, 12, 1)
]

start_dates = [
    datetime(2024, 8, 5), datetime(2024, 8, 18), datetime(2024, 8, 18),
    datetime(2024, 8, 19), datetime(2024, 10, 15), datetime(2024, 10, 15),
    datetime(2024, 10, 15), datetime(2024, 10, 20), datetime(2024, 10, 20),
    datetime(2024, 10, 15), datetime(2024, 10, 20), datetime(2024, 10, 20)
]