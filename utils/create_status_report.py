import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell import MergedCell

from config import black_border, yellow_fill, STATUS, plan_dates, start_dates, green_fill, red_fill
from database.main import SessionLocal
from database.models.Apartment import Apartment
from utils.calculate_percentage import calculate_completion_percentage


def generate_apartment_status_report(report_filename: str):
    session = SessionLocal()

    apartments_3 = session.query(Apartment).filter(Apartment.entrance_number == 3).all()
    apartments_4 = session.query(Apartment).filter(Apartment.entrance_number == 4).all()
    total_apartments = session.query(Apartment).all()

    completion_percentage_3 = calculate_completion_percentage(apartments_3)
    completion_percentage_4 = calculate_completion_percentage(apartments_4)
    total_percentage = calculate_completion_percentage(total_apartments)

    try:
        with open('completion_data.json', 'r') as f:
            completion_data = json.load(f)
    except FileNotFoundError:
        completion_data = []

    completion_dict = {
        entry['status_id']: entry['completion_date'] for entry in completion_data
    }

    data = []
    for idx, status in enumerate(STATUS):
        plan_date = plan_dates[idx] if idx < len(plan_dates) else None
        formatted_plan_date = plan_date.strftime("%Y-%m-%d") if plan_date else ''
        start_date = start_dates[idx] if idx < len(start_dates) else None
        formatted_start_date= start_date.strftime("%Y-%m-%d") if start_date else ''

        status_id = idx
        completion_date = completion_dict.get(status_id, None)

        if completion_date:
            formatted_completion_date = completion_date
        else:
            formatted_completion_date = "в работе"

        data.append({
            "Статус": f"{status}\r\n",
            " ": '\t                 \t               \t',
            "3 подъезд": f"{completion_percentage_3.get(status, 0):.2f}%",
            "4 подъезд": f"{completion_percentage_4.get(status, 0):.2f}%",
            "Общий  \t": f"{total_percentage.get(status, 0):.2f}%",
            "Дата начало/завершения": f'{formatted_start_date} - {formatted_completion_date}',
            "План": formatted_plan_date,
            "Факт": f"{completion_date}" if completion_date else ""
        })

    df = pd.DataFrame(data)

    df.to_excel(report_filename, index=False, engine='openpyxl')

    wb = load_workbook(report_filename)
    ws = wb.active
    ws.title = "Отчет по этапам"

    ws.insert_rows(1)
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Отчет по статусам квартир'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    for row in ws.iter_rows(min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.endswith('%'):
                try:
                    percentage_value = float(cell.value.rstrip('%'))

                    left_cell = ws.cell(row=cell.row, column=cell.column - 3)

                    if percentage_value == 100:
                        left_cell.fill = green_fill
                        left_cell.border = black_border
                    elif percentage_value >= 1:
                        left_cell.fill = red_fill
                        left_cell.border = black_border
                    else:
                        left_cell.fill = yellow_fill
                        left_cell.border = black_border
                except ValueError:
                    pass

    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        if column_letter:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(report_filename)

    return report_filename
