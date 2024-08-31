from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell import MergedCell

from config import green_fill, red_fill, DEADLINE, black_border
from database.main import SessionLocal
from database.models.Apartment import Apartment
from database.models.ApartmentStage import StageEnum
from utils.days_until_deadline import days_until_deadline
from utils.format_deadline import format_deadline_date


def generate_apartment_stage_report(report_filename: str):
    session = SessionLocal()
    apartments = session.query(Apartment).all()

    data = []
    for apartment in apartments:
        first_stage = next((stage for stage in apartment.stages if stage.stage == StageEnum.FIRST), None)
        second_stage = next((stage for stage in apartment.stages if stage.stage == StageEnum.SECOND), None)
        status = apartment.status if apartment.status is not None else ""
        status_date = apartment.status_date if apartment.status_date is not None else ""

        data.append({
            "Номер квартиры": apartment.number,
            "Статус": f"{status}\r\n{status_date}",
            "Дата начало работ": apartment.start_date,
            "Первый этап готов": 'Принят' if (first_stage and first_stage.is_finished) else 'Нет',
            "Причина отказа от 1 этапа": first_stage.comment,
            "Второй этап готов": 'Принят' if (second_stage and second_stage.is_finished) else 'Нет',
            "Причина отказа от 2 этапа": second_stage.comment,
            "Дата завершения работ": apartment.end_date,
            "Дедлайн": days_until_deadline(),
            "Чистота производства": apartment.clear_level,
            "АВР": 'Подписан' if apartment.is_accepted else 'Не подписан'
        })

    df = pd.DataFrame(data)

    df.to_excel(report_filename, index=False, engine='openpyxl')

    wb = load_workbook(report_filename)
    ws = wb.active
    ws.title = "Отчет по этапам"

    ws.insert_rows(1)
    ws.merge_cells('A1:K1')
    ws['A1'] = 'Отчет по готовности этапов квартир'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    deadline = datetime.strptime(DEADLINE, '%Y-%m-%d')

    for row in ws.iter_rows(min_col=8, max_col=8):
        for cell in row:
            if cell.value is None:
                cell_date = datetime.today()
            elif isinstance(cell.value, datetime):
                cell_date = cell.value
            else:
                continue

            if cell_date <= deadline:
                cell.fill = green_fill
                cell.border = black_border
            else:
                cell.fill = red_fill
                cell.border = black_border

    for row in ws.iter_rows(min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        if column_letter:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(report_filename)

    return report_filename
