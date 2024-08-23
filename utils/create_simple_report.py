from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.cell import MergedCell

from config import DEADLINE
from database.main import SessionLocal
from database.models.Apartment import Apartment
from database.models.ApartmentStage import StageEnum
from utils.check_repairing import count_apartments_with_conditions
from utils.days_until_deadline import days_until_deadline
from utils.format_deadline import format_deadline_date


def generate_apartment_stage_simple_report(report_filename: str):
    session = SessionLocal()

    apartments = session.query(Apartment).all()

    total_first_stage_finished = 0
    total_second_stage_finished = 0
    total_both_stages_finished = 0

    data = []
    for apartment in apartments:
        first_stage = next((stage for stage in apartment.stages if stage.stage == StageEnum.FIRST), None)
        second_stage = next((stage for stage in apartment.stages if stage.stage == StageEnum.SECOND), None)

        first_stage_finished = first_stage.is_finished if first_stage else False
        second_stage_finished = second_stage.is_finished if second_stage else False

        if first_stage_finished:
            total_first_stage_finished += 1
        if second_stage_finished:
            total_second_stage_finished += 1
        if first_stage_finished and second_stage_finished:
            total_both_stages_finished += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по этапам"

    ws['A1'] = 'Отчет по готовности этапов квартир'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')

    ws['A2'] = 'Этап'
    ws['B2'] = 'Количество'
    ws['A9'] = 'ГПР'
    ws['A8'] = 'В ремонте'
    ws['B8'] = count_apartments_with_conditions()
    ws['B9'] = format_deadline_date()
    ws['B10'] = days_until_deadline()

    for idx, row_data in enumerate(data, start=3):
        ws[f'A{idx}'] = row_data["Номер квартиры"]
        ws[f'B{idx}'] = row_data["Первый этап готов"]
        ws[f'C{idx}'] = row_data["Второй этап готов"]

    summary_start_row = len(data) + 4
    ws[f'A{summary_start_row}'] = 'Итоги:'
    ws[f'A{summary_start_row}'].font = Font(size=14, bold=True)

    ws[f'A{summary_start_row + 1}'] = '1 этап принят'
    ws[f'B{summary_start_row + 1}'] = total_first_stage_finished

    ws[f'A{summary_start_row + 2}'] = '2 этап принят'
    ws[f'B{summary_start_row + 2}'] = total_second_stage_finished

    ws[f'A{summary_start_row + 3}'] = 'Полностью завершено квартир'
    ws[f'B{summary_start_row + 3}'] = total_both_stages_finished

    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        if column_letter:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(report_filename)

    return report_filename
